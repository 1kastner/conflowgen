"""
The command line tool allows to create an XLSX file with all input distributions
which the user can change and then read in again to create synthetic container flows.
"""
import datetime
import os
import argparse
import typing

import openpyxl.worksheet.worksheet
import openpyxl.workbook
import openpyxl.styles

from conflowgen import metadata
from conflowgen.api.container_flow_generation_manager import ContainerFlowGenerationManager
from conflowgen.api.database_chooser import DatabaseChooser
from conflowgen.api.container_length_distribution_manager import ContainerLengthDistributionManager
from conflowgen.api.export_container_flow_manager import ExportContainerFlowManager

DEFAULT_ROW_OFFSET_OF_TABLE = 4


def _create_sheet(
        workbook: openpyxl.workbook.Workbook, title: str, index: typing.Optional[int] = None
) -> openpyxl.worksheet.worksheet.Worksheet:
    sheet = workbook.create_sheet(title, index)
    sheet["B2"] = title
    sheet["B2"].font = openpyxl.styles.Font(bold=True)
    sheet["B2"].alignment = openpyxl.styles.Alignment(horizontal="center")
    sheet.merge_cells('B2:C2')
    return sheet


def create_input_spreadsheet(file_path: str, overwrite: bool) -> None:
    """
    file_path: The path to the XLSX file to create
    """
    print(f"Create XLSX file at {file_path}")
    if os.path.isfile(file_path):
        if not overwrite:
            print(f"The file {file_path} already exists!")
            return
        print(f"The file {file_path} will be overwritten")

    workbook = openpyxl.Workbook()

    metadata_sheet = _create_sheet(workbook, "ConFlowGen metadata", 0)
    metadata_sheet["B4"] = "Version"
    metadata_sheet["C4"] = metadata.__version__
    metadata_sheet["B5"] = "Description"
    metadata_sheet["C5"] = metadata.__description__

    DatabaseChooser().create_new_sqlite_database(":memory:")

    container_flow_generation_sheet = _create_sheet(workbook, "Container flow generation")
    container_flow_generation_properties = ContainerFlowGenerationManager().get_properties()
    for i, (key, value) in enumerate(container_flow_generation_properties.items()):
        row_index = i + DEFAULT_ROW_OFFSET_OF_TABLE
        container_flow_generation_sheet["B" + str(row_index)] = key
        container_flow_generation_sheet["C" + str(row_index)] = value

    length_distribution_sheet = _create_sheet(workbook, "Container Length Distribution")
    container_length_distribution = ContainerLengthDistributionManager().get_container_length_distribution()
    for i, (container_length, percentage) in enumerate(container_length_distribution.items()):
        row_index = i + DEFAULT_ROW_OFFSET_OF_TABLE
        length_distribution_sheet["B" + str(row_index)] = str(container_length)
        length_distribution_sheet["C" + str(row_index)] = percentage

    del workbook["Sheet"]  # the default sheet we haven't used
    for sheet in workbook:
        for column in "ABCDEFG":
            sheet.column_dimensions[column].width = 20

    workbook.active = metadata_sheet
    workbook.save(file_path)


def create_container_flow(file_path: str, overwrite: bool):
    """
    file_path: The path to the XLSX file to read in
    """
    print(f"Read XLSX file at {file_path} to create container flows")

    if not os.path.isfile(file_path):
        print(f"The file {file_path} does not exists!")
        return

    workbook = openpyxl.load_workbook(file_path)

    file_path_without_ending = ".".join(file_path.split(".")[:-1])

    path_to_sqlite_file = os.path.join(
        os.path.dirname(file_path),
        file_path_without_ending + ".sqlite"
    )
    if os.path.isfile(path_to_sqlite_file):
        if not overwrite:
            print(f"The file {path_to_sqlite_file} already exists!")
            return
        print(f"The file {path_to_sqlite_file} will be overwritten")

    DatabaseChooser().create_new_sqlite_database(
        path_to_sqlite_file,
        overwrite=overwrite
    )
    print(f"Create container flow database: {path_to_sqlite_file}")
    container_flow_generation_manager = ContainerFlowGenerationManager()

    container_flow_generation_sheet = workbook["Container flow generation"]
    container_flow_generation_properties = {}
    for row_index in range(DEFAULT_ROW_OFFSET_OF_TABLE, container_flow_generation_sheet.max_row + 1):
        key = container_flow_generation_sheet["B" + str(row_index)].value
        value = container_flow_generation_sheet["C" + str(row_index)].value
        if key in [
            "conflowgen_version"
        ]:  # some information can be ignored
            continue
        container_flow_generation_properties[key] = value
    print(f"Parsed container flow generation properties: {container_flow_generation_properties}")
    if not container_flow_generation_properties["end_date"]:
        container_flow_generation_properties["end_date"] = datetime.datetime.now()
    if not container_flow_generation_properties["start_date"]:
        container_flow_generation_properties["start_date"] = (
                container_flow_generation_properties["end_date"] - datetime.timedelta(days=21)
        )
    print(f"Updated container flow generation properties: {container_flow_generation_properties} (start and end date "
          "must be set)")
    container_flow_generation_manager.set_properties(**container_flow_generation_properties)
    print(f"Set container flow generation properties: {container_flow_generation_manager.get_properties()}")

    length_distribution_sheet = workbook["Container Length Distribution"]
    container_length_distribution = {}
    for row_index in range(DEFAULT_ROW_OFFSET_OF_TABLE, length_distribution_sheet.max_row + 1):
        key = length_distribution_sheet["B" + str(row_index)].value
        value = length_distribution_sheet["C" + str(row_index)].value
        container_length_distribution[key] = value
    print(f"Parsed container length distribution: {container_length_distribution}")
    container_length_distribution_manager = ContainerLengthDistributionManager()
    container_length_distribution_manager.set_container_length_distribution(container_length_distribution)
    print(f"Set container length distribution: "
          f"{container_length_distribution_manager.get_container_length_distribution()}")

    print("Generating container flow data...")
    container_flow_generation_manager.generate()

    print("Exporting container flow data to CSV...")
    path_to_export_folder = os.path.dirname(file_path_without_ending)
    ExportContainerFlowManager().export(
        folder_name=file_path_without_ending + "_CSV_Export",
        path_to_export_folder=path_to_export_folder,
        overwrite=overwrite,
    )
    print("Creating container flow data has finished!")


def command_line_interface():
    parser = argparse.ArgumentParser(
        prog="ConFlowGen Spreadsheet Interface",
        description=(
            "The command line tool allows to create an XLSX file with all input distributions "
            "which the user can change and then read in again to create synthetic container flows."
        ),
    )
    valid_actions = [
        "create_input_spreadsheet",
        "create_container_flow"
    ]
    parser.add_argument(
        "action",
        choices=valid_actions,
        help="The action to perform",
    )
    parser.add_argument(
        "filepath",
        help="The path to the XLSX file",
    )
    parser.add_argument(
        "--overwrite",
        help="Whether existing files shall be overwritten",
        default=False,
        action="store_true",
    )
    args = parser.parse_args()
    filepath = os.path.abspath(os.path.join(
        os.curdir,
        args.filepath
    ))
    if args.action == "create_input_spreadsheet":
        create_input_spreadsheet(filepath, args.overwrite)
    elif args.action == "create_container_flow":
        create_container_flow(filepath, args.overwrite)


if __name__ == "__main__":
    command_line_interface()
